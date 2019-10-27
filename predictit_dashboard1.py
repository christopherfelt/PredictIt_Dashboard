#Excel Packages
import openpyxl
from openpyxl.styles import Font

#PredictIt JSON Packages
import requests
import json
import pprint
from datetime import datetime
from datetime import timedelta

#Get Market Data
response = requests.get('https://www.predictit.org/api/marketdata/all/')
market_json_raw = response.text
market_json = json.loads(market_json_raw)
markets = market_json['markets']

#Get Markets with less than ten days to end date
#TODO: Make in to function to find any market based on days to close. Put in seperate file
#This will return list

def getMarketByRemainingDuration(days):

    lessThanTen = []
    add_i = False
    for i in range(0,len(markets)):
        for key in markets[i]:
            if key == 'contracts':
                contracts = markets[i][key]
                for con in contracts:
                    con_dict = con
                    for item in con_dict.keys():
                        if item == "dateEnd":
                            if con_dict[item] != "N/A":
                                end_date = datetime.strptime(con_dict[item][:10], "%Y-%m-%d")
                                now = datetime.now()
                                difference = end_date-now
                                if difference < timedelta(days=days) :
                                    add_i = True
        if add_i:
            lessThanTen.append(i)
            add_i = False

    return lessThanTen

# five_day_list = getMarketByRemainingDuration(30)
#
# for i in five_day_list:
#     print("Markets with only five days left: ", i)

#Markets based on twitter
#TODO: Possible Function to find markets based on key word. It will return list

def getMarketByKeyword(keyword):

    tweet_markets = []
    add_j = False
    for j in range(0,len(markets)):
        for key in markets[j]:
            if key == 'name':
                if keyword in markets[j][key]:
                    add_j = True
        if add_j:
            tweet_markets.append(j)
            add_j = False

    return tweet_markets

twitter_markets = getMarketByKeyword('tweets')

for i in twitter_markets:
    print('Twitter Markets: ', i)

# print("Twitter Markets: ", len(tweet_markets))

#Get market and contract data
#TODO: Fix so it goes through all twitter markets
#TODO: Make into function that takes in a list and returns a dictionary
excel_dict = {}
excel_all_contracts = {}
excel_contract_dict = {}
market_list = []

# i=150

print("Number of Markets: " + str(len(markets)))
for tw in twitter_markets:
    for key in markets[tw]:
        #market level information: id, name, short name, image, url, timeStamp, status Open
        if key == 'id':
            excel_dict['id'] = markets[tw][key]
        if key == 'name':
            excel_dict['name'] = markets[tw][key]

        if key == 'contracts':
            contracts = markets[tw][key]
            for con in contracts:
                #individual contract dictionary/information:
                # id, dateend, image, name, long name, short name, status,
                # lasttradeprice, bestbuyyescost, bestbuynocost, bestsellyescost, bestsellnocost, lastcloseprice
                # displayorder
                con_dict = con
                for item in con_dict.keys():
                    if item == 'id':
                        excel_contract_dict['id'] = con_dict[item]
                    elif item == 'name':
                        excel_contract_dict['name'] = con_dict[item]
                    elif item == 'displayOrder':
                        excel_contract_dict['displayorder'] = con_dict[item]
                    elif item == "dateEnd":
                        end_date = datetime.strptime(con_dict[item][:10], "%Y-%m-%d").date()
                        end_time = datetime.strptime(con_dict[item][-8:], "%H:%M:%S").time()
                        end_datetime = datetime.combine(end_date,end_time)
                        now = datetime.now().date()
                        difference = end_date-now
                        # print('Difference: ', difference)
                        excel_contract_dict['enddate'] = end_datetime
                    elif item == "lastTradePrice":
                        excel_contract_dict['current_price'] = con_dict[item]
                    elif item == "bestBuyYesCost":
                        excel_contract_dict['bestbuyyescost'] = con_dict[item]
                    elif item == "bestBuyNoCost":
                        excel_contract_dict['bestbuynocost'] = con_dict[item]
                    elif item == "bestBuyYesCost":
                        excel_contract_dict['bestbuyyescost'] = con_dict[item]
                    elif item == "bestSellYesCost":
                        excel_contract_dict['bestsellyescost'] = con_dict[item]
                    elif item == "bestSellNoCost":
                        excel_contract_dict['bestsellnocost'] = con_dict[item]
                    elif item == "lastClosePrice":
                        excel_contract_dict['lastcloseprice'] = con_dict[item]
                    else:
                        pass
                excel_all_contracts[excel_contract_dict['name']] = excel_contract_dict
                excel_contract_dict = {}
            excel_dict['contracts'] = excel_all_contracts
            excel_all_contracts = {}
        else:
            pass
    market_list.append(excel_dict)
    excel_dict = {}


wb = openpyxl.Workbook()

sheet_list = wb.sheetnames

#going to need another dictionary and for loop to wrap around the if

#if not do set up

excel_counter = 0

for m in market_list:

    excel_dict = m

    if excel_dict['id'] not in sheet_list:
        ws = wb.create_sheet(str(excel_dict['id']), excel_counter)
        # Setup
        ws['a3'] = 'Start Date'
        ws['a4'] = 'End Date'
        ws['a7'] = 'Time Left'
        ws['b6'] = 'Days'
        ws['c6'] = 'Hours'
        ws['d6'] = 'Minutes'
        ws['a9'] = 'Contract'
        ws['b9'] = 'Current Price'
        ws['c9'] = 'Best Buy Yes Cost'
        ws['d9'] = 'Best Buy No Cost'
        ws['e9'] = 'Best Sell Yes Cost'
        ws['f9'] = 'Best Sell No Cost'
        ws['g9'] = 'Last Close Price'

        cell_list = ['a1','a3', 'a4', 'a7', 'a9', 'b6', 'b9', 'c6', 'c9', 'd6',
                     'd9','e9', 'f9', 'g9']

        for cell in cell_list:
            format_cell = ws[cell]
            format_cell.font = Font(bold=True)

        ws['a1'] = excel_dict['name']

        excel_contracts_column_dict = excel_dict['contracts']

        row = 10
        column = 1

        for contract, contract_dict in excel_contracts_column_dict.items():
            for key, value in contract_dict.items():
                if key not in ['id', 'displayorder','enddate']:
                    if value == None:
                        ws.cell(row=row, column=column, value = "n/a")
                    else:
                        ws.cell(row=row, column=column, value=value)
                    column = column + 1
            row = row + 1
            column = 1

    excel_counter = excel_counter+1

print(excel_dict)
wb.save('Test_Workbook7.xlsx')
